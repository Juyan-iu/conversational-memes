#!/usr/bin/env python3
"""
Build an Excel workbook for annotating MEMECONV stance labels and auditing visual descriptions.

Example on the target server
----------------------------
cd /home/exouser/git/conversational-memes/07_analysis

python -m venv audit_env
source audit_env/bin/activate
python -m pip install -r label_audit/requirements.txt

python label_audit/make_label_audit_workbook.py \
  --input ../03_filter_and_label/labeled_final/labeled_memes.jsonl \
  --image-root ../03_filter_and_label/labeled_final \
  --output ./label_audit/meme_label_audit_100.xlsx \
  --sample-size 100 \
  --seed 42 \
  --stratify-by-month

If image paths in `downloaded_images` are relative to a different folder, pass:

python make_label_audit_workbook.py \
  --input /path/to/labeled_memes.jsonl \
  --image-root /path/to/labeled_final \
  --output ./meme_label_audit_100.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable
from urllib.error import URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PILImage


STANCE_KEYS = ("sarcastic", "humorous", "offensive")

REVIEW_COLUMNS: list[tuple[str, int]] = [
    ("sample_id", 9),
    ("uid", 24),
    ("month", 11),
    ("conversation_context", 58),
    ("meme_reply_text", 36),
    ("meme_image_url", 32),
    ("meme_image", 36),
    ("image_note", 24),
    ("annotate_sarcastic", 18),
    ("annotate_humorous", 18),
    ("annotate_offensive", 18),
    ("stance_notes", 32),
    ("visual_description", 58),
    ("vd_caption_excluded", 20),
    ("vd_single_sentence", 18),
    ("vd_visual_elements_1_3", 21),
    ("vd_metaphor_1_3", 18),
    ("vd_self_contained_1_3", 23),
    ("vd_faithfulness_1_3", 22),
    ("vd_context_independent", 23),
    ("vd_notes", 34),
    ("exclude_item", 13),
    ("reviewer", 16),
    ("review_date", 16),
]

CODEBOOK_ROWS: list[tuple[str, str, str, str]] = [
    (
        "Purpose",
        "Annotate stance labels and audit the existing visual description.",
        "Stance is annotated from scratch. Visual description is evaluated against the existing generated text.",
        "",
    ),
    (
        "Shown materials",
        "Conversation context, meme image, meme reply text, and existing visual description.",
        "Judge the meme_reply as used in context.",
        "",
    ),
    (
        "Sarcastic",
        "Binary stance annotation.",
        "True when the utterance is sarcastic or ironic, expresses the opposite of its literal meaning, or mocks someone/something.",
        "Annotate Yes/No directly. Do not compare against an existing stance label.",
    ),
    (
        "Humorous",
        "Binary stance annotation.",
        "True when the utterance is humorous/funny, uses comedy, jokes, playfulness, or recognizable meme humor.",
        "Annotate Yes/No directly. Do not compare against an existing stance label.",
    ),
    (
        "Offensive",
        "Binary stance annotation.",
        "True when the utterance is offensive/aggressive, attacks, demeans, or uses hateful language toward someone/something.",
        "Annotate Yes/No directly. Do not compare against an existing stance label.",
    ),
    (
        "VD Caption Excluded",
        "Yes/No",
        "The visual description must not quote or directly mention overlaid caption text.",
        "No if caption leakage is visible.",
    ),
    (
        "VD Single Sentence",
        "Yes/No",
        "The description should be one concise sentence.",
        "No for lists, multi-sentence outputs, or commentary.",
    ),
    (
        "VD Visual Elements",
        "1-3",
        "1 = missing/wrong key characters, objects, expressions, setting, or action; 2 = partly right but incomplete; 3 = salient visual elements captured accurately.",
        "",
    ),
    (
        "VD Metaphor / Symbolic Meaning",
        "1-3",
        "1 = no or wrong metaphor; 2 = plausible but generic/shallow; 3 = clear symbolic meaning, or appropriately avoids over-interpretation when no metaphor is visible.",
        "",
    ),
    (
        "VD Self-Contained",
        "1-3",
        "1 = hard to imagine/recreate the image; 2 = rough scene only; 3 = enough detail to recreate the meme image without seeing it.",
        "",
    ),
    (
        "VD Image Faithfulness",
        "1-3",
        "1 = major hallucination or contradiction; 2 = minor errors/over-interpretation; 3 = faithful with no important errors.",
        "",
    ),
    (
        "VD Context Independent",
        "Yes/No",
        "The description should stand on the image itself rather than relying on thread-specific context.",
        "",
    ),
    (
        "Sources",
        "MEMECONV Appendix A.4.2; MEMECONV Appendix C.2.3; MemeCap; Memotion/SemEval.",
        "Criteria adapt prompt requirements plus meme evaluation dimensions: visual clarity/faithfulness, metaphor grounding, caption exclusion, stance attributes.",
        "",
    ),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a 100-item Excel workbook for MEMECONV label auditing."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to labeled_memes.jsonl, a JSON file, or a directory containing labeled_memes.jsonl or records/*.json.",
    )
    parser.add_argument(
        "--output",
        default="meme_label_audit_100.xlsx",
        help="Output .xlsx path.",
    )
    parser.add_argument("--sample-size", type=int, default=100)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--image-root",
        default=None,
        help="Base folder for relative image paths, usually the labeled_final output directory.",
    )
    parser.add_argument(
        "--stratify-by-month",
        action="store_true",
        help="Sample as evenly as possible by month when month information is available.",
    )
    parser.add_argument(
        "--download-missing",
        action="store_true",
        help="Download meme images from stored URLs when local files are missing.",
    )
    parser.add_argument(
        "--download-dir",
        default=None,
        help="Directory for downloaded images when --download-missing is used.",
    )
    parser.add_argument(
        "--stance-output",
        default=None,
        help=(
            "Optional CSV path for the sampled records' existing stance labels "
            "(sample_id, uid, month, existing_sarcastic, existing_humorous, existing_offensive)."
        ),
    )
    parser.add_argument(
        "--include-existing-stance",
        action="store_true",
        help=(
            "Append existing stance-label columns to the workbook. Leave this off for blind annotation."
        ),
    )
    return parser.parse_args()


def read_json_file(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL at {path}:{line_no}: {exc}") from exc
            if isinstance(item, dict):
                yield item


def load_records(input_path: Path) -> list[dict[str, Any]]:
    if input_path.is_file():
        if input_path.suffix.lower() == ".jsonl":
            return list(iter_jsonl(input_path))
        if input_path.suffix.lower() == ".json":
            data = read_json_file(input_path)
            if isinstance(data, list):
                return [x for x in data if isinstance(x, dict)]
            if isinstance(data, dict):
                for key in ("records", "items", "data"):
                    if isinstance(data.get(key), list):
                        return [x for x in data[key] if isinstance(x, dict)]
                return [data]
        raise ValueError(f"Unsupported input file type: {input_path}")

    if not input_path.is_dir():
        raise FileNotFoundError(input_path)

    preferred = input_path / "labeled_memes.jsonl"
    if preferred.exists():
        return list(iter_jsonl(preferred))

    jsonl_files = sorted(input_path.glob("*.jsonl"))
    if jsonl_files:
        records: list[dict[str, Any]] = []
        for path in jsonl_files:
            records.extend(iter_jsonl(path))
        return records

    records_dir = input_path / "records"
    if records_dir.exists():
        json_files = sorted(records_dir.glob("*.json"))
    else:
        json_files = sorted(input_path.glob("*.json"))

    records = []
    for path in json_files:
        data = read_json_file(path)
        if isinstance(data, dict):
            records.append(data)
    return records


def first_text(*values: Any) -> str:
    for value in values:
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def get_post_text(post: Any) -> str:
    if not isinstance(post, dict):
        return ""
    record = post.get("record") if isinstance(post.get("record"), dict) else {}
    return first_text(
        post.get("text"),
        post.get("display_text"),
        post.get("caption"),
        record.get("text"),
    )


def safe_cell(value: Any, limit: int = 32000) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)
    text = re.sub(r"\s+\n", "\n", text).strip()
    return text[:limit]


def get_nested(data: dict[str, Any], *keys: str, default: Any = None) -> Any:
    cur: Any = data
    for key in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key)
    return default if cur is None else cur


def bool_to_yes_no(value: Any) -> str:
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)) and value in {0, 1}:
        return "Yes" if bool(value) else "No"
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"yes", "y", "true", "t", "1"}:
            return "Yes"
        if text in {"no", "n", "false", "f", "0"}:
            return "No"
    return ""


def extract_existing_stance(record: dict[str, Any]) -> dict[str, str]:
    """Return original stance labels as Yes/No strings when available."""
    candidates = [
        get_nested(record, "discourse_labels", "meme_reply", "stance", default={}),
        get_nested(record, "discourse_labels", "meme_reply", "stance_labels", default={}),
        get_nested(record, "annotation", "stance", default={}),
        get_nested(record, "meme_reply", "stance", default={}),
        record.get("stance_labels") or {},
        record.get("stance") or {},
    ]

    stance: dict[str, str] = {key: "" for key in STANCE_KEYS}
    for candidate in candidates:
        if not isinstance(candidate, dict):
            continue
        for key in STANCE_KEYS:
            value = bool_to_yes_no(candidate.get(key))
            if value and not stance[key]:
                stance[key] = value
    return stance


def extract_labels(record: dict[str, Any]) -> dict[str, str]:
    meme_labels = get_nested(record, "discourse_labels", "meme_reply", default={}) or {}
    visual = meme_labels.get("visual") or {}
    if not isinstance(visual, dict):
        visual = {}

    return {
        "visual_description": safe_cell(
            visual.get("visual_description")
            or record.get("visual_description")
            or get_nested(record, "annotation", "visual_description", default="")
        ),
    }


def build_context(record: dict[str, Any]) -> str:
    parts: list[str] = []

    def add(label: str, post: Any) -> None:
        text = get_post_text(post)
        if text:
            parts.append(f"[{label}] {text}")

    add("Original Post", record.get("original_post"))
    add("Quoted Post", record.get("quoted_post"))

    ancestors = record.get("ancestor_chain") or record.get("thread_ancestors") or []
    if isinstance(ancestors, list):
        for i, ancestor in enumerate(ancestors, 1):
            add(f"Reply {i}", ancestor)

    add("Parent Reply", record.get("parent_reply"))
    add("Meme Reply", record.get("meme_reply"))

    if not parts:
        return "(no text context available)"
    return "\n".join(parts)


def find_month(record: dict[str, Any]) -> str:
    candidates = [
        record.get("process_date"),
        record.get("labeled_at"),
        record.get("processed_at"),
        get_nested(record, "meme_reply", "created_at", default=None),
        get_nested(record, "meme_reply", "indexed_at", default=None),
    ]
    for value in candidates:
        if not value:
            continue
        match = re.search(r"(\d{4})-(\d{2})", str(value))
        if match:
            return f"{match.group(1)}-{match.group(2)}"
    return "unknown"


def eligible_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    kept = []
    for record in records:
        if not isinstance(record.get("meme_reply"), dict):
            continue
        labels = extract_labels(record)
        if not any(labels.values()):
            continue
        kept.append(record)
    return kept


def sample_records(
    records: list[dict[str, Any]],
    sample_size: int,
    seed: int,
    stratify_by_month: bool,
) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    records = list(records)
    if len(records) <= sample_size:
        rng.shuffle(records)
        return records

    if not stratify_by_month:
        return rng.sample(records, sample_size)

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        groups[find_month(record)].append(record)

    for group in groups.values():
        rng.shuffle(group)

    months = sorted(groups)
    quota = sample_size // len(months)
    remainder = sample_size % len(months)
    chosen: list[dict[str, Any]] = []
    leftovers: list[dict[str, Any]] = []

    for i, month in enumerate(months):
        target = quota + (1 if i < remainder else 0)
        group = groups[month]
        chosen.extend(group[:target])
        leftovers.extend(group[target:])

    if len(chosen) < sample_size:
        rng.shuffle(leftovers)
        chosen.extend(leftovers[: sample_size - len(chosen)])

    rng.shuffle(chosen)
    return chosen[:sample_size]


def image_items(record: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    downloaded = get_nested(record, "downloaded_images", "meme_reply", default=[]) or []
    if isinstance(downloaded, list):
        items.extend(x for x in downloaded if isinstance(x, dict))

    raw_images = get_nested(record, "meme_reply", "images", default=[]) or []
    if isinstance(raw_images, list):
        items.extend(x for x in raw_images if isinstance(x, dict))

    return items


def add_candidate_path(
    candidates: list[Path],
    value: Any,
    roots: Iterable[Path],
) -> None:
    if not isinstance(value, str) or not value.strip():
        return
    path = Path(value).expanduser()
    if path.is_absolute():
        candidates.append(path)
        return
    for root in roots:
        candidates.append(root / path)


def infer_extension(url: str, content_type: str) -> str:
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return suffix
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    return ".jpg"


def download_image(url: str, out_dir: Path, uid: str) -> Path | None:
    out_dir.mkdir(parents=True, exist_ok=True)
    try:
        req = Request(url, headers={"User-Agent": "MemeLabelAudit/1.0"})
        with urlopen(req, timeout=20) as response:
            data = response.read()
            content_type = response.headers.get("Content-Type", "")
    except (OSError, URLError):
        return None

    ext = infer_extension(url, content_type)
    out_path = out_dir / f"{uid or 'meme'}{ext}"
    out_path.write_bytes(data)
    return out_path


def resolve_meme_image(
    record: dict[str, Any],
    base_dir: Path,
    image_root: Path,
    download_missing: bool,
    download_dir: Path,
) -> tuple[Path | None, str, str]:
    uid = safe_cell(record.get("uid") or get_nested(record, "meme_reply", "uid", default=""))
    roots = [image_root, base_dir]
    candidates: list[Path] = []
    urls: list[str] = []

    for item in image_items(record):
        for key in ("local_path", "path", "file", "filename"):
            add_candidate_path(candidates, item.get(key), roots)
        url = item.get("url") or item.get("source_url") or item.get("fullsize") or item.get("thumb")
        if isinstance(url, str) and url.strip():
            urls.append(url.strip())

    if uid:
        for root in roots:
            meme_dir = root / "images" / "meme_reply" / uid
            if meme_dir.exists():
                candidates.extend(
                    path
                    for path in sorted(meme_dir.iterdir())
                    if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".gif"}
                )

    for path in candidates:
        if path.exists() and path.is_file():
            return path, (urls[0] if urls else ""), "local"

    if download_missing and urls:
        downloaded = download_image(urls[0], download_dir, uid)
        if downloaded and downloaded.exists():
            return downloaded, urls[0], "downloaded"

    note = "missing local image"
    if not urls:
        note = "missing local image and URL"
    return None, (urls[0] if urls else ""), note


def add_validation(ws, cell_range: str, values: str, prompt: str) -> None:
    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
    dv.prompt = prompt
    dv.error = f"Choose one of: {values}"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def add_thumbnail(ws, image_path: Path, anchor: str, tmp_dir: Path, max_size: tuple[int, int]) -> str:
    try:
        with PILImage.open(image_path) as img:
            img.thumbnail(max_size)
            if img.mode not in {"RGB", "RGBA"}:
                img = img.convert("RGBA")
            thumb_path = tmp_dir / f"{anchor}_{image_path.stem}.png"
            img.save(thumb_path)
            xl_img = XLImage(str(thumb_path))
            xl_img.width, xl_img.height = img.size
            ws.add_image(xl_img, anchor)
            return f"embedded {img.size[0]}x{img.size[1]}"
    except Exception as exc:  # noqa: BLE001
        return f"image embed failed: {exc}"


def write_review_sheet(
    wb: Workbook,
    records: list[dict[str, Any]],
    base_dir: Path,
    image_root: Path,
    download_missing: bool,
    download_dir: Path,
    thumb_dir: Path,
    include_existing_stance: bool,
) -> None:
    ws = wb.active
    ws.title = "Review_100"
    review_columns = list(REVIEW_COLUMNS)
    if include_existing_stance:
        review_columns.extend(
            [
                ("existing_sarcastic", 18),
                ("existing_humorous", 18),
                ("existing_offensive", 18),
            ]
        )
    ws.append([name for name, _ in review_columns])

    for idx, (_, width) in enumerate(review_columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for sample_id, record in enumerate(records, 1):
        labels = extract_labels(record)
        stance = extract_existing_stance(record)
        image_path, image_url, image_note = resolve_meme_image(
            record,
            base_dir=base_dir,
            image_root=image_root,
            download_missing=download_missing,
            download_dir=download_dir,
        )

        row = [
            sample_id,
            safe_cell(record.get("uid") or get_nested(record, "meme_reply", "uid", default="")),
            find_month(record),
            build_context(record),
            get_post_text(record.get("meme_reply")),
            image_url,
            "",
            image_note,
            "",
            "",
            "",
            "",
            labels["visual_description"],
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        if include_existing_stance:
            row.extend([stance["sarcastic"], stance["humorous"], stance["offensive"]])
        ws.append(row)
        excel_row = ws.max_row
        ws.row_dimensions[excel_row].height = 170

        if image_url:
            ws.cell(excel_row, 6).hyperlink = image_url
            ws.cell(excel_row, 6).style = "Hyperlink"

        if image_path:
            note = add_thumbnail(ws, image_path, f"G{excel_row}", thumb_dir, (250, 210))
            ws.cell(excel_row, 8).value = note
        elif image_note:
            ws.cell(excel_row, 8).value = image_note

    style_sheet(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    max_row = ws.max_row
    add_validation(ws, f"I2:K{max_row}", "Yes,No", "Annotate this stance label directly.")
    add_validation(ws, f"N2:O{max_row}", "Yes,No", "Yes/No according to the visual-description codebook.")
    add_validation(ws, f"P2:S{max_row}", "1,2,3", "1=poor, 2=partial, 3=good.")
    add_validation(ws, f"T2:T{max_row}", "Yes,No", "Is the description context-independent?")
    add_validation(ws, f"V2:V{max_row}", "Yes,No", "Exclude this item from analysis?")

    ws["I1"].comment = Comment("Annotate whether the meme reply is sarcastic/ironic/mocking.", "Codex")
    ws["J1"].comment = Comment("Annotate whether the meme reply is humorous/funny/playful.", "Codex")
    ws["K1"].comment = Comment("Annotate whether the meme reply is offensive/aggressive/demeaning.", "Codex")
    ws["M1"].comment = Comment("Existing visual_description generated during labeling.", "Codex")
    ws["P1"].comment = Comment("1=wrong/missing, 2=partial, 3=salient visual elements captured.", "Codex")
    ws["Q1"].comment = Comment("1=wrong/missing metaphor, 2=generic, 3=well-grounded or avoids over-interpretation.", "Codex")
    ws["R1"].comment = Comment("1=not reconstructable, 2=partly reconstructable, 3=self-contained.", "Codex")
    ws["S1"].comment = Comment("1=major hallucination, 2=minor issues, 3=faithful.", "Codex")
    if include_existing_stance:
        ws["Y1"].comment = Comment("Existing sarcastic label from the source dataset.", "Codex")
        ws["Z1"].comment = Comment("Existing humorous label from the source dataset.", "Codex")
        ws["AA1"].comment = Comment("Existing offensive label from the source dataset.", "Codex")


def write_codebook_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Codebook")
    ws.append(["Item", "Scale", "Rule", "Notes"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 45
    for row in CODEBOOK_ROWS:
        ws.append(row)
    style_sheet(ws)
    ws.freeze_panes = "A2"


def write_stance_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "sample_id",
        "uid",
        "month",
        "existing_sarcastic",
        "existing_humorous",
        "existing_offensive",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for sample_id, record in enumerate(records, 1):
            stance = extract_existing_stance(record)
            writer.writerow(
                {
                    "sample_id": sample_id,
                    "uid": safe_cell(record.get("uid") or get_nested(record, "meme_reply", "uid", default="")),
                    "month": find_month(record),
                    "existing_sarcastic": stance["sarcastic"],
                    "existing_humorous": stance["humorous"],
                    "existing_offensive": stance["offensive"],
                }
            )


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    base_dir = input_path.parent if input_path.is_file() else input_path
    image_root = Path(args.image_root).expanduser().resolve() if args.image_root else base_dir
    download_dir = (
        Path(args.download_dir).expanduser().resolve()
        if args.download_dir
        else output_path.parent / "downloaded_audit_images"
    )

    records = load_records(input_path)
    records = eligible_records(records)
    if not records:
        raise SystemExit("No eligible labeled meme records found.")

    sampled = sample_records(
        records,
        sample_size=args.sample_size,
        seed=args.seed,
        stratify_by_month=args.stratify_by_month,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    with tempfile.TemporaryDirectory(prefix="meme_audit_thumbs_") as tmp:
        write_review_sheet(
            wb,
            sampled,
            base_dir=base_dir,
            image_root=image_root,
            download_missing=args.download_missing,
            download_dir=download_dir,
            thumb_dir=Path(tmp),
            include_existing_stance=args.include_existing_stance,
        )
        write_codebook_sheet(wb)
        wb.save(output_path)

    if args.stance_output:
        stance_output = Path(args.stance_output).expanduser().resolve()
        write_stance_csv(sampled, stance_output)
        print(f"Wrote stance labels: {stance_output}")

    print(f"Wrote {output_path}")
    print(f"Records loaded: {len(records)}")
    print(f"Records sampled: {len(sampled)}")


if __name__ == "__main__":
    main()
