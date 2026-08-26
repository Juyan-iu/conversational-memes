#!/usr/bin/env python3
"""
Bluesky-style thread card -> screenshot -> Excel

Usage:
  python make_annotation_xlsx.py \
    --input ./labeled_final/labeled_memes.jsonl \
    --output ./annotation_200.xlsx \
    --n 200 --seed 42
"""

import json, random, argparse, io, tempfile, time
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


# ── utils ────────────────────────────────────────────────────────
def get_month(r):
    val = (r.get("original_post") or {}).get("created_at", "")
    if val:
        try:
            ts = float(val)
            return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m")
        except Exception:
            pass
    return "unknown"

def get_images(node):
    if not node:
        return []
    return [
        img.get("source_url") or img.get("url", "")
        for img in (node.get("images") or [])
        if img.get("source_url") or img.get("url", "")
    ]

def get_meme_img_url(r):
    imgs = get_images(r.get("meme_reply") or {})
    return imgs[0] if imgs else ""

def get_llm_labels(r):
    labels = r.get("discourse_labels", {}) or {}
    meme   = labels.get("meme_reply") or {}
    stance = meme.get("stance") or {}
    return {
        "discourse": meme.get("discourse_function", ""),
        "sarcastic":  "Yes" if stance.get("sarcastic") else "No",
        "humorous":   "Yes" if stance.get("humorous")  else "No",
        "offensive":  "Yes" if stance.get("offensive") else "No",
    }

def esc(t):
    return (t or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")


# ── HTML helpers ─────────────────────────────────────────────────
def imgs_html(urls):
    if not urls:
        return ""
    tags = "".join(
        f'<img src="{esc(u)}" onerror="this.style.display=\'none\'">'
        for u in urls
    )
    return f'<div class="img-row">{tags}</div>'

def post_block(node, role, indent=0, show_connector=False):
    """한 포스트 블록 렌더링 (handle 없음)"""
    if not node:
        return ""
    text = (node.get("text") or "").strip()
    imgs = get_images(node)

    # 텍스트도 이미지도 없으면 스킵
    if not text and not imgs:
        return ""

    role_cfg = {
        "post":     ("#EBF5FB", "#1565C0", "📌 Original Post"),
        "ancestor": ("#F5F5F5", "#555555", "💬"),
        "parent":   ("#E8F8F5", "#1B5E20", "↩ Parent Reply"),
        "meme":     ("#FEF9E7", "#7D6608", "🎭 Meme Reply"),
    }
    bg, label_color, label = role_cfg.get(role, ("#FFF", "#000", ""))

    indent_px = indent * 20
    connector = f'<div class="connector" style="margin-left:{indent_px + 10}px;"></div>' if show_connector else ""

    return f"""
    {connector}
    <div class="post-block" style="background:{bg};margin-left:{indent_px}px;">
      <div class="role-label" style="color:{label_color};">{label}</div>
      {f'<div class="post-text">{esc(text)}</div>' if text else ""}
      {imgs_html(imgs)}
    </div>"""

def quoted_block(node):
    if not node:
        return ""
    text = (node.get("text") or "").strip()
    imgs = get_images(node)
    if not text and not imgs:
        return ""
    return f"""
    <div class="quoted-embed">
      <div class="quoted-label">🔁 Quoted Post</div>
      {f'<div class="post-text">{esc(text)}</div>' if text else ""}
      {imgs_html(imgs)}
    </div>"""


# ── Full card HTML ────────────────────────────────────────────────
def make_card_html(record, no):
    month   = get_month(record)
    t_label = (record.get("thread_structure") or {}).get("label", "reply")
    orig    = record.get("original_post") or {}
    parent  = record.get("parent_reply") or {}
    meme    = record.get("meme_reply") or {}
    quoted  = record.get("quoted_post")
    chain   = record.get("ancestor_chain") or []

    blocks = []

    # 1. Original Post (+ quoted embed)
    orig_text = (orig.get("text") or "").strip()
    orig_imgs = get_images(orig)
    orig_html = f"""
    <div class="post-block" style="background:#EBF5FB;">
      <div class="role-label" style="color:#1565C0;">📌 Original Post</div>
      {f'<div class="post-text">{esc(orig_text)}</div>' if orig_text else ""}
      {imgs_html(orig_imgs)}
      {quoted_block(quoted)}
    </div>"""
    blocks.append(orig_html)

    # 2. Ancestor chain (re-reply일 때 중간 댓글)
    for i, anc in enumerate(chain):
        b = post_block(anc, "ancestor", indent=i+1, show_connector=True)
        if b:
            blocks.append(b)

    # 3. Parent reply (re-reply일 때)
    if t_label == "re-reply" and parent:
        indent = len(chain) + 1
        b = post_block(parent, "parent", indent=indent, show_connector=True)
        if b:
            blocks.append(b)

    # 4. Meme reply
    if t_label == "re-reply":
        meme_indent = len(chain) + (2 if parent else 1)
    else:
        meme_indent = 1
    blocks.append(post_block(meme, "meme", indent=meme_indent, show_connector=True))

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{
    font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;
    font-size:13px; background:#fff; width:900px; padding:12px;
  }}
  .card {{ border:2px solid #1B4F72; border-radius:10px; overflow:hidden; }}
  .card-header {{
    background:#1B4F72; color:#fff; padding:8px 14px;
    font-size:12px; display:flex; gap:10px; align-items:center;
  }}
  .badge {{
    background:rgba(255,255,255,0.22); padding:2px 8px;
    border-radius:10px; font-size:11px;
  }}
  .thread {{ padding:10px 12px 14px; display:flex; flex-direction:column; }}
  .connector {{
    width:2px; height:12px; background:#C0C0C0; flex-shrink:0;
  }}
  .post-block {{
    border:1px solid #DDD; border-radius:8px;
    padding:10px 12px; margin-top:2px;
  }}
  .role-label {{
    font-size:11px; font-weight:bold; margin-bottom:5px;
  }}
  .post-text {{
    color:#222; line-height:1.6; white-space:pre-wrap;
    word-break:break-word; font-size:13px;
  }}
  .img-row {{ display:flex; flex-wrap:wrap; gap:6px; margin-top:8px; }}
  .img-row img {{
    max-width:560px; max-height:440px; object-fit:contain;
    border-radius:6px; border:1px solid #DDD;
  }}
  .quoted-embed {{
    border:1px solid #CCC; border-radius:6px;
    padding:8px 10px; margin-top:8px; background:#F9F9F9;
  }}
  .quoted-label {{
    font-size:11px; color:#888; font-weight:bold; margin-bottom:4px;
  }}
</style></head><body>
<div class="card">
  <div class="card-header">
    <span style="font-weight:bold;">#{no}</span>
    <span class="badge">📅 {esc(month)}</span>
    <span class="badge">{esc(t_label)}</span>
  </div>
  <div class="thread">
    {"".join(blocks)}
  </div>
</div>
</body></html>"""


# ── screenshot ───────────────────────────────────────────────────
def screenshot_card(page, html, tmp_path):
    tmp_html = tmp_path / "card.html"
    tmp_html.write_text(html, encoding="utf-8")
    page.goto(f"file://{tmp_html}")
    page.wait_for_load_state("networkidle", timeout=12000)
    time.sleep(1.5)
    card = page.query_selector(".card")
    return card.screenshot() if card else page.screenshot()


# ── Excel styles ─────────────────────────────────────────────────
def fill(c):   return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style="thin", color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def add_anchored_image(ws, png_bytes, col, row):
    img = XLImage(io.BytesIO(png_bytes))
    m1  = AnchorMarker(col=col-1, colOff=0, row=row-1, rowOff=0)
    m2  = AnchorMarker(col=col,   colOff=0, row=row,   rowOff=0)
    img.anchor = TwoCellAnchor(_from=m1, to=m2)
    ws.add_image(img)


# ── Excel ────────────────────────────────────────────────────────
def create_xlsx(records, output_path):
    wb = Workbook()

    # Instructions sheet
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i.column_dimensions["A"].width = 100
    lines = [
        ("밈 댓글 담화 기능 & Stance 라벨링 가이드", True, 14, "1B4F72", "FFFFFF"),
        ("", False, 10, None, "000000"),
        ("■ 라벨링 방법", True, 12, None, "1B4F72"),
        ("1. Annotation 시트에서 각 행의 카드 이미지를 확인하세요.", False, 10, None, "000000"),
        ("2. 카드는 블루스카이 화면처럼 포스트→댓글→밈 순서로 표시됩니다.", False, 10, None, "000000"),
        ("3. 본인 열에 라벨을 입력하세요. 다른 어노테이터 결과를 보지 말고 독립적으로 판단하세요.", False, 10, None, "000000"),
        ("4. Exclude 열: 해당 항목을 제외해야 한다고 판단되면 Yes 입력 (연구자 전용).", False, 10, None, "000000"),
        ("", False, 10, None, "000000"),
        ("■ 카드 구조", True, 12, None, "1B4F72"),
        ("  📌 Original Post — 원 포스트 (파란 배경)", False, 10, None, "000000"),
        ("  🔁 Quoted Post   — 인용된 포스트 (원 포스트 안에 표시)", False, 10, None, "000000"),
        ("  💬 (회색)        — 중간 댓글 (re-reply인 경우)", False, 10, None, "000000"),
        ("  ↩ Parent Reply   — 부모 댓글 (re-reply인 경우, 초록 배경)", False, 10, None, "000000"),
        ("  🎭 Meme Reply    — 라벨링 대상 밈 댓글 (노란 배경)", False, 10, None, "000000"),
        ("", False, 10, None, "000000"),
        ("■ 담화 기능 입력값", True, 12, None, "1B4F72"),
        ("  React.Respond   — 이전 발화에 직접 반응 (동의/반대/정보). 대화 마무리 느낌.", False, 10, None, "000000"),
        ("  React.Rejoinder — 감정 표현/되묻기/코멘트. 대화가 이어질 것 같은 느낌.", False, 10, None, "000000"),
        ("  Open.Give / Open.Demand / Open.Command / Open.Attend — 새 대화 시작 (드물음)", False, 10, None, "000000"),
        ("", False, 10, None, "000000"),
        ("■ Stance 입력값 (각각 Yes / No)", True, 12, None, "1B4F72"),
        ("  Sarcastic — 비꼬거나 반어적인 톤", False, 10, None, "000000"),
        ("  Humorous  — 웃음을 유발하는 유쾌한 톤", False, 10, None, "000000"),
        ("  Offensive — 공격적이거나 혐오적인 내용", False, 10, None, "000000"),
        ("", False, 10, None, "000000"),
        ("■ 주의사항", True, 12, None, "1B4F72"),
        ("  - 판단이 어려우면 Note 칸에 이유를 기록해주세요.", False, 10, None, "000000"),
    ]
    for ri, (text, bold, size, bg, fg) in enumerate(lines, 1):
        c = ws_i.cell(row=ri, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, size=size, color=fg)
        c.alignment = Alignment(wrap_text=True)
        if bg:
            c.fill = fill(bg)

    # Annotation sheet (Meme URL 열 제거)
    ws = wb.create_sheet("Annotation")
    headers = [
        "No", "Card", "Exclude?",
        "[LLM] Discourse","[LLM] Sarcastic","[LLM] Humorous","[LLM] Offensive",
        "[Ann1] Discourse","[Ann1] Sarcastic","[Ann1] Humorous","[Ann1] Offensive","[Ann1] Note",
        "[Ann2] Discourse","[Ann2] Sarcastic","[Ann2] Humorous","[Ann2] Offensive","[Ann2] Note",
        "[Ann3] Discourse","[Ann3] Sarcastic","[Ann3] Humorous","[Ann3] Offensive","[Ann3] Note",
    ]
    col_widths = [
        5, 100, 10,
        20,13,13,13,
        20,13,13,13,28,
        20,13,13,13,28,
        20,13,13,13,28,
    ]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = fill("1B4F72")
        cell.alignment = aln("center","center")
        cell.border    = bdr()

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # LLM 열 숨기기 (D~G)
    for col in ["D","E","F","G"]:
        ws.column_dimensions[col].hidden = True

    ROW_H = 420  # pt 더 넉넉하게

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={"width": 950, "height": 1200})

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            for i, record in enumerate(records, 2):
                no = i - 1
                print(f"  [{no}/{len(records)}] processing...", end="\r", flush=True)

                llm = get_llm_labels(record)

                try:
                    html = make_card_html(record, no)
                    png  = screenshot_card(page, html, tmp_path)
                    add_anchored_image(ws, png, col=2, row=i)
                except Exception as e:
                    print(f"\n  [WARN] #{no}: {e}")

                ws.row_dimensions[i].height = ROW_H

                row_data = [
                    no, "", "",
                    llm["discourse"],llm["sarcastic"],llm["humorous"],llm["offensive"],
                    "","","","","",
                    "","","","","",
                    "","","","","",
                ]
                bgs = [
                    None, None, "FEF9E7",
                    "F0F0F0","F0F0F0","F0F0F0","F0F0F0",
                    "FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF",
                    "EAF4FB","EAF4FB","EAF4FB","EAF4FB","EAF4FB",
                    "F0FFF0","F0FFF0","F0FFF0","F0FFF0","F0FFF0",
                ]
                for ci, (val, bg) in enumerate(zip(row_data, bgs), 1):
                    if ci == 2:
                        continue
                    cell = ws.cell(row=i, column=ci, value=val)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = bdr()
                    cell.alignment = aln()
                    if bg:
                        cell.fill = fill(bg)

        browser.close()

    wb.save(output_path)
    print(f"\n[DONE] {output_path} ({len(records)} records)")


# ── main ─────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default="./labeled_final/labeled_memes.jsonl")
    parser.add_argument("--output", default="./annotation_200.xlsx")
    parser.add_argument("--n",      type=int, default=200)
    parser.add_argument("--seed",   type=int, default=42)
    args = parser.parse_args()

    random.seed(args.seed)

    records = []
    with open(args.input, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except Exception:
                    pass
    print(f"[LOAD] {len(records):,} records")

    sampled = random.sample(records, min(args.n, len(records)))
    print(f"[SAMPLE] {len(sampled)} records")
    print(f"[START] Rendering cards...")

    create_xlsx(sampled, args.output)


if __name__ == "__main__":
    main()
